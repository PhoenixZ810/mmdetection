import re
import argparse
import os
import csv
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

parser = argparse.ArgumentParser(description="log_name")
parser.add_argument("method", type=str, help='method name, used in csv/xlsx header')
parser.add_argument("epoch", type=int, help = 'train_epoch, uesd for checking whether training completed')
parser.add_argument('--work_dirs', type=str, default='work_dirs/', required=False, help='directory for saving results')
args = parser.parse_args()

def main():
    num = 0
    fail_num = 0
    none_exist_num = 0
    fail=[]
    none_exist=[]
    for dataset in sorted(os.listdir("rf100/")):
        print(f'\ndataset={dataset}, index={num}')

        # determine whether the dataset directory exists
        try:
            dirs = [os.path.join(args.work_dirs, dataset, d) for d in os.listdir(os.path.join(args.work_dirs, dataset)) 
                    if os.path.isdir(os.path.join(args.work_dirs, dataset, d))]
            num+=1
        except:
            print(f"{dataset} directory doesn't exist!")
            none_exist_num+=1
            none_exist.append(dataset)
            continue

        dirs.sort(key=os.path.getmtime)
        latest_dir = dirs[-1]
        latest_log_name = latest_dir.split('/')[-1]
        print('time='+latest_log_name)

        latest_log = latest_dir+f'/{latest_log_name}.log'  # get latest log name
        print(latest_log)

        with open(latest_log, 'r') as f:
            log = f.read()
        with open('rf100/'+dataset+'/train/_annotations.coco.json','r') as f:
            image=json.load(f)
            num_train = len(image['images'])  # get number of train images        
        with open('rf100/'+dataset+'/valid/_annotations.coco.json','r') as f:
            image=json.load(f)
            num_valid = len(image['images'])  # get number of valid images
        with open('scripts/labels_names.json') as f:
            label=json.load(f)
            for index in label:
                if index['name'] == dataset:
                    category = index['category']  # get category of dataset
        complete_flag=re.findall(r'Epoch\(val\) \[{}\]\[\d+/\d+\]'.format(args.epoch), log)  # find log of args.epoch's validing process
        
        # Check whether the training is complete
        if not complete_flag:
            fail_num+=1
            fail.append(dataset)
            print("------------------------------------------------------------------------------------")
            print(f'{dataset} train failed!')
            print(f'{fail_num} dataset failed!')
            print("------------------------------------------------------------------------------------")
            key_value=[dataset, category, num_train, num_valid, '', '', '', '', '']
        else:
            '''match result'''
            match_all = re.findall(r'The best checkpoint with ([\d.]+) coco/bbox_mAP at ([\d.]+) epoch', log)
            if match_all:
                match = match_all[-1]
                best_epoch = match[-1]
                print(f'best_epoch={best_epoch}')
                match_AP = re.findall(r'\[{}\]\[\d+/\d+\]    coco/bbox_mAP: (-?\d+\.?\d*)  coco/bbox_mAP_50: (-?\d+\.?\d*)  coco/bbox_mAP_75: -?\d+\.?\d*  coco/bbox_mAP_s: (-?\d+\.?\d*)  coco/bbox_mAP_m: (-?\d+\.?\d*)  coco/bbox_mAP_l: (-?\d+\.?\d*)'.format(best_epoch), log)
                print(f'match_AP={match_AP}')
                
                key_value = [dataset, category, num_train, num_valid]
                key_value.extend(match_AP[0])
            else:
                print("------------------------------------------------------------------------------------")
                print('log has no result!')
                print("------------------------------------------------------------------------------------")
                key_value=[dataset, category, num_train, num_valid, '', '', '', '', '']

        if num==1:
            result_csv = os.path.join(args.work_dirs,f'{latest_log_name}_final_eval.csv')
            print(result_csv)
            with open(result_csv, mode='a') as f:
                writer = csv.writer(f)
                header1 = ['Dataset', 'Category', 'Images', 'Images', args.method, args.method, args.method, args.method, args.method]
                writer.writerow(header1)
                header2 = ['', '', 'train', 'valid', 'mAP', 'mAP50', 'mAP_s', 'mAP_m', 'mAP_l']
                writer.writerow(header2)
                writer.writerow(key_value)

        else:
            with open(result_csv, mode='a') as f:
                writer = csv.writer(f)
                writer.writerow(key_value)

    # Convert .csv file to .xlsx file
    df = pd.read_csv(result_csv)
    result_xlsx = '{}.xlsx'.format(result_csv.split('.')[0])
    print(f'\n{result_xlsx} created!\n')
    df.to_excel(result_xlsx)
    wb = load_workbook(result_xlsx)
    ws = wb.active
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:J1')
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(result_xlsx)
    print(f'{none_exist_num} datasets were not trained:\n{none_exist}\n')
    print(f'{fail_num} training failed:\n{fail}')

if __name__ == "__main__":
    main()
     